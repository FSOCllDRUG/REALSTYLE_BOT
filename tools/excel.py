import json

import aiofiles
import openpyxl


async def json_to_excel(json_file, excel_file):
    async with aiofiles.open('tools/' + json_file, 'r', encoding='utf-8') as f:
        data = json.loads(await f.read())

    wb = openpyxl.Workbook()
    ws = wb.active

    # Переводим ключи на русский язык
    translate = {
        "insurance": "Страховка",
        "rate_BYN": "Курс BYN",
        "rate_RUB": "Курс RUB",
        "sneakers": "Кроссовки",
        "boots": "Ботинки",
        "slippers": "Шлёпанцы",
        "tops": "ВерхняяОдежда",
        "hoodies": "Худи",
        "pants": "Штаны",
        "tShirts": "Футболки",
        "shorts": "Шорты",
        "socks": "Носки",
        "panties": "НижнееБельё",
        "perfume": "Парфюм",
        "watches": "Часы",
        "sportBags": "СпортивныеСумки",
        "bigBags": "БольшиеСумки",
        "smallBags": "МаленькиеСумки",
        "hats": "КепкиШапки",
        "jewelry": "Украшения",
        "belts": "Ремни",
        "headphones": "Наушники",
        "keyboard": "Клавиатуры",
        "mouse": "Мыши",
        "computers": "КрупногабаритнаяТехника"
    }

    # Записываем заголовки в Excel
    ws.append(["Название", "BYN", "RUB", "% страховки"])

    # Записываем данные страховки и курсов
    ws.append(["Страховка", "", "", data["insurance"]])
    ws.append(["Курс", data["rate_BYN"], data["rate_RUB"], ""])

    # Записываем остальные данные
    items = {}
    for key, value in data.items():
        if key.startswith("ship_") or key.startswith("margin_"):
            parts = key.split('_')
            item = translate.get(parts[1], parts[1])
            type_ = "Доставка" if key.startswith("ship_") else "Комиссия"
            currency = parts[-1]

            if item not in items:
                items[item] = {"Доставка BYN": "", "Доставка RUB": "", "Комиссия BYN": "", "Комиссия RUB": ""}

            if type_ == "Доставка":
                items[item][f"Доставка {currency}"] = value
            elif type_ == "Комиссия":
                items[item][f"Комиссия {currency}"] = value

    for item, values in items.items():
        ws.append([f"Доставка {item}", values["Доставка BYN"], values["Доставка RUB"], ""])
        ws.append([f"Комиссия {item}", values["Комиссия BYN"], values["Комиссия RUB"], ""])

    # Сохраняем Excel файл
    wb.save(excel_file)


async def excel_to_json(excel_file, json_file):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    data = {}

    reverse_translate = {
        "Страховка": "insurance",
        "Курс": "rate",
        "Кроссовки": "sneakers",
        "Ботинки": "boots",
        "Шлёпанцы": "slippers",
        "ВерхняяОдежда": "tops",
        "Худи": "hoodies",
        "Штаны": "pants",
        "Футболки": "tShirts",
        "Шорты": "shorts",
        "Носки": "socks",
        "НижнееБельё": "panties",
        "Парфюм": "perfume",
        "Часы": "watches",
        "СпортивныеСумки": "sportBags",
        "БольшиеСумки": "bigBags",
        "МаленькиеСумки": "smallBags",
        "КепкиШапки": "hats",
        "Украшения": "jewelry",
        "Ремни": "belts",
        "Наушники": "headphones",
        "Клавиатуры": "keyboard",
        "Мыши": "mouse",
        "КрупногабаритнаяТехника": "computers"
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        item, byn, rub, insurance = row

        if item == "Страховка":
            data["insurance"] = insurance
        elif item == "Курс":
            if byn:
                data["rate_BYN"] = byn
            if rub:
                data["rate_RUB"] = rub
        else:
            parts = item.split()
            key_base = reverse_translate.get(parts[1], parts[1])
            type_ = parts[0]

            if type_ == "Доставка":
                if byn:
                    data[f'ship_{key_base}_BYN'] = byn
                if rub:
                    data[f'ship_{key_base}_RUB'] = rub
            elif type_ == "Комиссия":
                if byn:
                    data[f'margin_{key_base}_BYN'] = byn
                if rub:
                    data[f'margin_{key_base}_RUB'] = rub

    async with aiofiles.open('tools/' + json_file, 'w', encoding='utf-8') as f:
        await f.write(json.dumps(data, ensure_ascii=False, indent=4))
